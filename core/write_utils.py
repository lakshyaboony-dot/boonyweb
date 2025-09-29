import os
from openpyxl import Workbook, load_workbook

def save_write_text(user_id, mobile, current_day, write_text):
    filename = f"{user_id}_{mobile}.xlsx"
    headers = ["day", "write_assignment", "write_assessment"]

    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Progress"
        ws.append(headers)
        wb.save(filename)

    wb = load_workbook(filename)
    ws = wb.active
    ws.append([current_day, write_text, "Pending"])
    wb.save(filename)
